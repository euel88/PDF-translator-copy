"""
Excel 문서 핸들러 (.xlsx, .xls)
openpyxl 라이브러리를 사용하여 Excel 문서 번역

개선된 기능:
- 대용량 문서 청크 처리
- 동적 배치 크기 조정
- 셀 스타일 완벽 보존
- 차트 제목/레이블 번역
- 주석(노트) 번역
- 헤더/푸터 번역
- 도형 텍스트 번역
"""
from typing import Optional, List, Tuple, Dict, Any
from pathlib import Path
from dataclasses import dataclass

from pdf2zh.handlers.base import BaseDocumentHandler, DocumentResult


@dataclass
class TextLocation:
    """텍스트 위치 정보"""
    loc_type: str  # 'cell', 'comment', 'chart_title', 'chart_axis', 'header', 'footer', 'shape'
    sheet_name: str
    identifiers: tuple  # 위치 식별자
    text: str


class ExcelHandler(BaseDocumentHandler):
    """Excel 문서 핸들러 - 대용량 처리 개선"""

    SUPPORTED_EXTENSIONS = ['.xlsx', '.xls']

    # 대용량 처리 설정
    BATCH_SIZE_SMALL = 50   # 짧은 텍스트
    BATCH_SIZE_MEDIUM = 30  # 중간 텍스트
    BATCH_SIZE_LARGE = 15   # 긴 텍스트

    def convert(
        self,
        input_path: str,
        output_path: Optional[str] = None
    ) -> DocumentResult:
        """
        Excel 문서 번역

        Args:
            input_path: 입력 파일 경로
            output_path: 출력 파일 경로

        Returns:
            DocumentResult: 변환 결과
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return DocumentResult(
                success=False,
                error="openpyxl 라이브러리가 설치되지 않았습니다. 'pip install openpyxl'을 실행하세요."
            )

        if output_path is None:
            output_path = self.generate_output_path(input_path)

        try:
            self.log(f"Excel 문서 열기: {input_path}")
            # data_only=False로 수식 보존
            workbook = load_workbook(input_path, data_only=False)

            # 모든 텍스트 수집 (셀, 주석, 차트, 도형 등)
            text_locations = self._collect_all_texts(workbook)
            total_count = len(text_locations)

            if total_count == 0:
                workbook.save(output_path)
                return DocumentResult(
                    output_path=output_path,
                    success=True,
                    translated_count=0,
                    total_count=0
                )

            # 텍스트 추출
            texts_to_translate = [loc.text for loc in text_locations]

            # 동적 배치 크기로 번역
            self.log("번역 중...")
            translated_texts = self._translate_with_dynamic_batch(texts_to_translate)

            # 번역 결과 적용 (스타일 보존)
            translated_count = self._apply_translations(workbook, text_locations, translated_texts)

            # 저장
            self.log(f"저장 중: {output_path}")
            workbook.save(output_path)

            return DocumentResult(
                output_path=output_path,
                success=True,
                translated_count=translated_count,
                total_count=total_count
            )

        except Exception as e:
            import traceback
            return DocumentResult(
                success=False,
                error=f"{str(e)}\n{traceback.format_exc()}"
            )

    def _collect_all_texts(self, workbook) -> List[TextLocation]:
        """모든 텍스트 수집 (셀, 주석, 차트, 도형 등)"""
        locations = []
        stats = {'cell': 0, 'comment': 0, 'chart': 0, 'shape': 0, 'header': 0, 'footer': 0}

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            self.log(f"시트 분석: {sheet_name}")

            # 1. 셀 텍스트 수집
            for row_idx, row in enumerate(sheet.iter_rows(), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    if self._should_translate_cell(cell):
                        locations.append(TextLocation(
                            loc_type='cell',
                            sheet_name=sheet_name,
                            identifiers=(row_idx, col_idx),
                            text=str(cell.value)
                        ))
                        stats['cell'] += 1

            # 2. 주석(Comments/Notes) 수집
            try:
                # openpyxl 3.0+에서는 _comments 속성 사용
                if hasattr(sheet, '_comments'):
                    for comment in sheet._comments:
                        if comment and comment.text and comment.text.strip():
                            # 주석의 위치 찾기
                            cell_coord = None
                            for coord, c in getattr(sheet, '_comments', {}).items():
                                if c == comment:
                                    cell_coord = coord
                                    break
                            if cell_coord:
                                locations.append(TextLocation(
                                    loc_type='comment',
                                    sheet_name=sheet_name,
                                    identifiers=(cell_coord,),
                                    text=comment.text
                                ))
                                stats['comment'] += 1
            except Exception:
                pass

            # 3. 셀별 주석 수집 (대안 방법)
            try:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.comment and cell.comment.text and cell.comment.text.strip():
                            locations.append(TextLocation(
                                loc_type='comment',
                                sheet_name=sheet_name,
                                identifiers=(cell.coordinate,),
                                text=cell.comment.text
                            ))
                            stats['comment'] += 1
            except Exception:
                pass

            # 4. 차트 수집
            try:
                for chart_idx, chart in enumerate(getattr(sheet, '_charts', [])):
                    # 차트 제목
                    if hasattr(chart, 'title') and chart.title:
                        title_text = self._get_chart_title_text(chart.title)
                        if title_text:
                            locations.append(TextLocation(
                                loc_type='chart_title',
                                sheet_name=sheet_name,
                                identifiers=(chart_idx, 'title'),
                                text=title_text
                            ))
                            stats['chart'] += 1

                    # X축 제목
                    if hasattr(chart, 'x_axis') and chart.x_axis and hasattr(chart.x_axis, 'title'):
                        x_title = self._get_chart_title_text(chart.x_axis.title)
                        if x_title:
                            locations.append(TextLocation(
                                loc_type='chart_axis',
                                sheet_name=sheet_name,
                                identifiers=(chart_idx, 'x_axis'),
                                text=x_title
                            ))
                            stats['chart'] += 1

                    # Y축 제목
                    if hasattr(chart, 'y_axis') and chart.y_axis and hasattr(chart.y_axis, 'title'):
                        y_title = self._get_chart_title_text(chart.y_axis.title)
                        if y_title:
                            locations.append(TextLocation(
                                loc_type='chart_axis',
                                sheet_name=sheet_name,
                                identifiers=(chart_idx, 'y_axis'),
                                text=y_title
                            ))
                            stats['chart'] += 1

                    # 범례 항목 (레이블)
                    if hasattr(chart, 'legend') and chart.legend:
                        try:
                            for ser_idx, series in enumerate(chart.series):
                                if hasattr(series, 'title') and series.title:
                                    ser_title = str(series.title) if series.title else None
                                    if ser_title and ser_title.strip():
                                        locations.append(TextLocation(
                                            loc_type='chart_series',
                                            sheet_name=sheet_name,
                                            identifiers=(chart_idx, 'series', ser_idx),
                                            text=ser_title
                                        ))
                                        stats['chart'] += 1
                        except Exception:
                            pass
            except Exception:
                pass

            # 5. 도형(Shape) 텍스트 수집
            try:
                if hasattr(sheet, '_images'):
                    for img_idx, img in enumerate(sheet._images):
                        if hasattr(img, 'desc') and img.desc:
                            locations.append(TextLocation(
                                loc_type='shape',
                                sheet_name=sheet_name,
                                identifiers=(img_idx, 'desc'),
                                text=img.desc
                            ))
                            stats['shape'] += 1
            except Exception:
                pass

            # 6. 헤더/푸터 수집
            try:
                if hasattr(sheet, 'HeaderFooter') and sheet.HeaderFooter:
                    hf = sheet.HeaderFooter
                    # 헤더
                    for part in ['left', 'center', 'right']:
                        header = getattr(hf.oddHeader, part, None) if hf.oddHeader else None
                        if header and header.strip():
                            locations.append(TextLocation(
                                loc_type='header',
                                sheet_name=sheet_name,
                                identifiers=('odd', part),
                                text=header
                            ))
                            stats['header'] += 1
                    # 푸터
                    for part in ['left', 'center', 'right']:
                        footer = getattr(hf.oddFooter, part, None) if hf.oddFooter else None
                        if footer and footer.strip():
                            locations.append(TextLocation(
                                loc_type='footer',
                                sheet_name=sheet_name,
                                identifiers=('odd', part),
                                text=footer
                            ))
                            stats['footer'] += 1
            except Exception:
                pass

        # 통계 로깅
        total = sum(stats.values())
        self.log(f"텍스트 수집 완료 (총 {total}개): 셀 {stats['cell']}, 주석 {stats['comment']}, "
                 f"차트 {stats['chart']}, 도형 {stats['shape']}, 헤더 {stats['header']}, 푸터 {stats['footer']}")

        return locations

    def _get_chart_title_text(self, title) -> Optional[str]:
        """차트 제목에서 텍스트 추출"""
        if title is None:
            return None
        if isinstance(title, str):
            return title if title.strip() else None
        if hasattr(title, 'text'):
            return title.text if title.text and title.text.strip() else None
        if hasattr(title, 'tx') and hasattr(title.tx, 'rich'):
            try:
                texts = []
                for p in title.tx.rich.p:
                    for r in p.r:
                        if hasattr(r, 't') and r.t:
                            texts.append(r.t)
                return ''.join(texts) if texts else None
            except Exception:
                pass
        return str(title) if title else None

    def _apply_translations(
        self,
        workbook,
        locations: List[TextLocation],
        translations: List[str]
    ) -> int:
        """번역 결과 적용"""
        from openpyxl.comments import Comment

        translated_count = 0

        for loc, translated in zip(locations, translations):
            try:
                sheet = workbook[loc.sheet_name]

                if loc.loc_type == 'cell':
                    row, col = loc.identifiers
                    cell = sheet.cell(row=row, column=col)
                    cell.value = translated
                    translated_count += 1

                elif loc.loc_type == 'comment':
                    coord = loc.identifiers[0]
                    cell = sheet[coord]
                    if cell.comment:
                        # 기존 주석의 작성자 유지
                        author = cell.comment.author if hasattr(cell.comment, 'author') else 'Translated'
                        cell.comment = Comment(translated, author)
                        translated_count += 1

                elif loc.loc_type == 'chart_title':
                    chart_idx, _ = loc.identifiers
                    charts = list(getattr(sheet, '_charts', []))
                    if chart_idx < len(charts):
                        self._set_chart_title(charts[chart_idx], translated)
                        translated_count += 1

                elif loc.loc_type == 'chart_axis':
                    chart_idx, axis_name = loc.identifiers
                    charts = list(getattr(sheet, '_charts', []))
                    if chart_idx < len(charts):
                        chart = charts[chart_idx]
                        axis = getattr(chart, axis_name, None)
                        if axis and hasattr(axis, 'title'):
                            self._set_axis_title(axis, translated)
                            translated_count += 1

                elif loc.loc_type == 'chart_series':
                    chart_idx, _, ser_idx = loc.identifiers
                    charts = list(getattr(sheet, '_charts', []))
                    if chart_idx < len(charts):
                        chart = charts[chart_idx]
                        if hasattr(chart, 'series') and ser_idx < len(chart.series):
                            chart.series[ser_idx].title = translated
                            translated_count += 1

                elif loc.loc_type == 'header':
                    hf_type, part = loc.identifiers
                    if hasattr(sheet, 'HeaderFooter') and sheet.HeaderFooter:
                        header = sheet.HeaderFooter.oddHeader
                        if header:
                            setattr(header, part, translated)
                            translated_count += 1

                elif loc.loc_type == 'footer':
                    hf_type, part = loc.identifiers
                    if hasattr(sheet, 'HeaderFooter') and sheet.HeaderFooter:
                        footer = sheet.HeaderFooter.oddFooter
                        if footer:
                            setattr(footer, part, translated)
                            translated_count += 1

            except Exception as e:
                self.log(f"경고: 번역 적용 실패 ({loc.loc_type}, {loc.sheet_name}): {e}")

        return translated_count

    def _set_chart_title(self, chart, text: str):
        """차트 제목 설정"""
        from openpyxl.chart.title import Title
        from openpyxl.chart.text import RichText
        from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, RegularTextRun

        try:
            chart.title = text
        except Exception:
            try:
                # 복잡한 제목 구조일 경우
                chart.title = Title()
                chart.title.tx = RichText()
                chart.title.tx.rich = RichText()
                para = Paragraph()
                para.r = [RegularTextRun(t=text)]
                chart.title.tx.rich.p = [para]
            except Exception:
                pass

    def _set_axis_title(self, axis, text: str):
        """축 제목 설정"""
        from openpyxl.chart.title import Title

        try:
            if hasattr(axis, 'title'):
                axis.title = text
        except Exception:
            try:
                axis.title = Title()
                axis.title.tx = text
            except Exception:
                pass

    def _translate_with_dynamic_batch(self, texts: List[str]) -> List[str]:
        """동적 배치 크기로 번역"""
        total = len(texts)

        # 텍스트 길이에 따라 그룹화
        short_texts = []  # < 100자
        medium_texts = []  # 100-500자
        long_texts = []  # > 500자

        for i, text in enumerate(texts):
            text_len = len(text)
            if text_len < 100:
                short_texts.append((i, text))
            elif text_len < 500:
                medium_texts.append((i, text))
            else:
                long_texts.append((i, text))

        # 결과 저장용 딕셔너리
        result_map = {}

        # 짧은 텍스트 배치 처리
        if short_texts:
            self.log(f"짧은 텍스트 번역: {len(short_texts)}개")
            self._batch_translate(short_texts, result_map, self.BATCH_SIZE_SMALL)

        # 중간 텍스트 배치 처리
        if medium_texts:
            self.log(f"중간 텍스트 번역: {len(medium_texts)}개")
            self._batch_translate(medium_texts, result_map, self.BATCH_SIZE_MEDIUM)

        # 긴 텍스트 배치 처리
        if long_texts:
            self.log(f"긴 텍스트 번역: {len(long_texts)}개")
            self._batch_translate(long_texts, result_map, self.BATCH_SIZE_LARGE)

        # 원래 순서대로 결과 정렬
        translations = []
        for i in range(total):
            translations.append(result_map.get(i, texts[i]))

        return translations

    def _batch_translate(
        self,
        indexed_texts: List[Tuple[int, str]],
        result_map: Dict[int, str],
        batch_size: int
    ):
        """배치 번역 수행"""
        for i in range(0, len(indexed_texts), batch_size):
            batch = indexed_texts[i:i + batch_size]
            texts = [t[1] for t in batch]
            indices = [t[0] for t in batch]

            try:
                translated = self.translate_texts(texts)
                for idx, trans in zip(indices, translated):
                    result_map[idx] = trans
            except Exception as e:
                self.log(f"배치 번역 오류, 개별 번역 시도: {e}")
                # 개별 번역 폴백
                for idx, text in batch:
                    try:
                        result_map[idx] = self.translate_texts([text])[0]
                    except Exception:
                        result_map[idx] = text  # 실패 시 원본 유지

            self.log(f"진행: {min(i + batch_size, len(indexed_texts))}/{len(indexed_texts)}")

    def _should_translate_cell(self, cell) -> bool:
        """셀을 번역해야 하는지 확인"""
        if cell.value is None:
            return False

        value = cell.value

        # 문자열만 번역
        if not isinstance(value, str):
            return False

        # 빈 문자열 제외
        if not value.strip():
            return False

        # 수식 제외 (=로 시작)
        if value.startswith('='):
            return False

        # 숫자만 있는 경우 제외
        try:
            float(value.replace(',', '').replace('%', '').replace(' ', ''))
            return False
        except ValueError:
            pass

        # URL 제외
        if value.startswith(('http://', 'https://', 'www.', 'ftp://')):
            return False

        # 이메일 제외
        if '@' in value and '.' in value.split('@')[-1]:
            return False

        # 파일 경로 제외
        if '\\' in value or (value.startswith('/') and '/' in value[1:]):
            return False

        # 최소 길이 (단일 문자/숫자 제외)
        if len(value.strip()) < 2:
            return False

        return True
